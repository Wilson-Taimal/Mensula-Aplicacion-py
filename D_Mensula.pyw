from tkinter import *
ventana = Tk()
ventana.geometry("1190x990")
ventana.title("DISEÑO ESTRUCTURAL MÉNSULA")
ventana.iconbitmap('D:\\BIBLIOTECA PERSONAL\\Programación\\Python\\logo-wat.ico')

# Función para programar boton "Calcular"

def calcular():
    print (' ')
    print (' - DISEÑO ESTRUCTURAL MÉNSULA - ')
    # Propiedades de los materiales
    fc = float (txt21.get())/10
    fy = float (txt22.get())/10
    u = float (txt23.get())
    fiv = float (txt24.get())
    fif = float (txt25.get())
    Nb = float (txt26.get())
    Nbe = float (txt27.get())

    # Geometría            
    a = float (txt30.get())
    b = float (txt31.get())
    c = float (txt32.get())
    h = float (txt33.get())
    r = float (txt34.get())
    bw = float (txt35.get())
    d = h - r
    hmin = 0.5*d
    rtxt3.delete(0, 'end')
    rtxt3.insert (0, "{:.2f}".format(hmin))
    x = a/d
    if (x < 1):
        opc1 = 'Ok'
        rtxt2.delete(0, 'end')
        rtxt2.insert (0, opc1)
    else:
        opc2 = 'No cumple'
        rtxt2.delete(0, 'end')
        rtxt2.insert (0, opc2)

    # Cargas de diseño
    print ('')
    print ('Cargas de diseño')    
    Vu = float (txt02.get())
    print (" Vu  = %.2f" %Vu, '_ kN')

    Nuc1 = float (txt04.get())
    if (Nuc1 <= Vu):
        opc1 = 'Ok'
        rtxt4.delete(0, 'end')
        rtxt4.insert (0, opc1)
    else:
        opc2 = 'No cumple'
        rtxt4.delete(0, 'end')
        rtxt4.insert (0, opc2)

    Nuc2 = 0.2*Vu
    txt06.delete(0, 'end')
    txt06.insert (0, "{:.2f}".format(Nuc2))

    Nuc = max(Nuc1, Nuc2)
    txt08.delete(0, 'end')
    txt08.insert (0, "{:.2f}".format(Nuc))
    print (" Nuc = %.2f" %Nuc, '_ kN')

    Mu = (Vu*a) + Nuc*(h-d)
    txt10.delete(0, 'end')
    txt10.insert (0, "{:.1f}".format(Mu))
    print (" Mu  = %.2f" %Mu, '_ kN.cm')

    # Verificación Vu
    Vc1 = 0.2*fiv*fc*bw*d
    rtxt6.delete(0, 'end')
    rtxt6.insert (0, "{:.0f}".format(Vc1))

    Vc2 = ((3.3+0.08*fc*10)*fiv*bw*10*d*10)/1000
    rtxt7.delete(0, 'end')
    rtxt7.insert (0, "{:.0f}".format(Vc2))

    Vc3 = 11*fiv*bw*d
    rtxt8.delete(0, 'end')
    rtxt8.insert (0, "{:.0f}".format(Vc3))

    Vc = min(Vc1, Vc2, Vc3)
    rtxt9.delete(0, 'end')
    rtxt9.insert (0, "{:.0f}".format(Vc))
    print (" Vc. Resistencia del concreto %.2f" %Vc, '_ kN')

    if (Vu <= Vc):
        opc1 = 'Ok'
        rtxt10.delete(0, 'end')
        rtxt10.insert (0, opc1)
    else:
        opc2 = 'No cumple'
        rtxt10.delete(0, 'end')
        rtxt10.insert (0, opc2)

    # Cálculo de refuerzo. Refuerzo para resistir Vu
    Avf = Vu/(fiv * fy * u)
    rtxt22.delete(0, 'end')
    rtxt22.insert (0, "{:.2f}".format(Avf))

    # Refuerzo para resistir Nuc
    An = Nuc/(fif*fy)
    rtxt24.delete(0, 'end')
    rtxt24.insert (0, "{:.2f}".format(An))

    # Refuerzo para resistir Mu
    Af = Mu/(fif*fy*(d-a/2))
    rtxt25.delete(0, 'end')
    rtxt25.insert (0, "{:.2f}".format(Af))
    
    # Refuerzo principal
    print ('')
    print ('Cálculo de refuerzo')
    As1 = Af + An
    rtxt27.delete(0, 'end')
    rtxt27.insert (0, "{:.2f}".format(As1))

    As2 = 2/3*Avf + An
    rtxt28.delete(0, 'end')
    rtxt28.insert (0, "{:.2f}".format(As2))

    As3 = 0.04*(fc/fy)*bw*d
    rtxt29.delete(0, 'end')
    rtxt29.insert (0, "{:.2f}".format(As3))

    As = max(As1, As2, As3)  
    rtxt30.delete(0, 'end')
    rtxt30.insert (0, "{:.2f}".format(As))
    print (" As. Acero principal: %.2f" %As, '_ cm²')  

    import f_Asb as Asb
    Asb = Asb.f_Asb(Nb)
    import math
    Cant = math.ceil(As/Asb)
    s = round((bw-15)/(Cant-1),0)

    # Verificación refuerzo mínimo
    Asc = Asb * Cant
    rtxt12.delete(0, 'end')
    rtxt12.insert (0, "{:.2f}".format(Asc))
    print (" Asc. Acero colocado: %.02f" %Asc, '_ cm²')
    pcal = Asc/(bw*d)
    rtxt13.delete(0, 'end')
    rtxt13.insert (0, "{:.5f}".format(pcal))
    print (" pcal = %.5f" %pcal)

    pmin = 0.04*fc/fy    
    rtxt14.delete(0, 'end')
    rtxt14.insert (0, "{:.5f}".format(pmin))
    print (" pmin = %.5f" %pmin)
       
    if (pcal >= pmin):
        opc1 = 'Ok'
        rtxt15.delete(0, 'end')
        rtxt15.insert (0, opc1)
    else:
        opc2 = 'No cumple'   
        rtxt15.delete(0, 'end')
        rtxt15.insert (0, opc2)
    
    # Refuerzo en estribos
    Ah = 0.5*(As - An)   
    rtxt34.delete(0, 'end')
    rtxt34.insert (0, "{:.2f}".format(Ah))
    print (" Ah. Acero en estribo: %.02f" %Ah, '_ cm²') 

    import f_Asb as Asb
    Asbe = Asb.f_Asb(Nbe)
    import math
    Cante = math.ceil(Ah/Asbe)
    lde = 2/3*d    
    rtxt33.delete(0, 'end')
    rtxt33.insert (0, "{:.2f}".format(lde))
    se = round((lde/Cante), 0)

    # Refuerzo auxiliar
    Aa = 0.002*bw*c    
    import f_Asb as Asb
    AsbAa = Asb.f_Asb(Nb)
    import math
    CantAa = math.ceil(Aa/AsbAa)
    sAa = round((bw-15)/(CantAa-1),0)
    print (" Aa. Acero auxiliar: %.02f" %Aa, '_ cm²')    

    # Distribución refuerzo
    print ('') 
    print ('Distribución de refuerzo')
    Reffp = str(Cant) + ' barras N° ' +  str("{:.0f}".format(Nb)) + " C/ " + "{:.0f}".format(s) + ' cm'    
    dtxt1.delete(0, 'end')
    dtxt1.insert (0, Reffp) 
    print (' Ref. Principal:', Cant, 'barras N°', "{:.0f}".format(Nb), "C/ %.0f" %s, ' cm') 
    
    Reffe = str(Cante) + ' estribos N° ' +  str("{:.0f}".format(Nbe)) + " C/ " + "{:.0f}".format(se) + ' cm'    
    dtxt2.delete(0, 'end')
    dtxt2.insert (0, Reffe)
    print (' Ref. Estribos:', Cante, 'barras N°', "{:.0f}".format(Nbe), "C/ %.0f" %se, ' cm')

    Reffa = str(CantAa) + ' barras N° ' +  str("{:.0f}".format(Nb)) + " C/ " + "{:.0f}".format(sAa) + ' cm'    
    dtxt3.delete(0, 'end')
    dtxt3.insert (0, Reffa)
    print (' Ref. Auxiliar:', CantAa, 'barras N°', "{:.0f}".format(Nb), "C/ %.0f" %sAa, ' cm')  

    # Detalles soldadura
    print ('') 
    print ('Detalle de soldadura')    
    import f_db as db
    db = db.f_db (Nb)
    dtxt4.delete(0, 'end')
    dtxt4.insert (0, "{:.2f}".format(db)) 
    print (" db. Diámetro refuerzo principal: %.2f" %db, '_ cm')
       
    lw = 3/4*db
    dtxt5.delete(0, 'end')
    dtxt5.insert (0, "{:.2f}".format(lw)) 
    print (" lw. Longitud de la soldadura: %.2f" %lw, '_ cm')

    tw = 1/2*db
    dtxt6.delete(0, 'end')
    dtxt6.insert (0, "{:.2f}".format(tw)) 
    print (" tw. Espesor de la soldadura: %.2f" %tw, '_ cm')

    print ('')
    print ('Diseño terminado con exito')
    
# Función para programar boton "Guardar xls"

def reporte():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    book = Workbook()
    #book = load_workbook('D:\BIBLIOTECA PERSONAL\Programación\Python\Mensula\PlantillaMensula.xlsx')
    book = load_workbook('D:\IEB_MENSULA\PlantillaMensula.xlsx')
    sheet = book.active

     # Materiales
    fc = float (txt21.get())/10
    sheet['C6'] = fc*10
    sheet['C6'].font = Font('Arial', size=10)

    fy = float (txt22.get())/10
    sheet['C7'] = fy*10
    sheet['C7'].font = Font('Arial', size=10)

    u = float (txt23.get())
    sheet['C8'] = u
    sheet['C8'].font = Font('Arial', size=10) 

    fiv = float (txt24.get())
    sheet['C9'] = fiv
    sheet['C9'].font = Font('Arial', size=10)

    fif = float (txt25.get())
    sheet['C10'] = fif
    sheet['C10'].font = Font('Arial', size=10)

    Nb = float (txt26.get())
    sheet['C11'] = Nb
    sheet['C11'].font = Font('Arial', size=10)

    Nbe = float (txt27.get())
    sheet['C12'] = Nbe
    sheet['C12'].font = Font('Arial', size=10)

    sheet['C13'] = Nb
    sheet['C13'].font = Font('Arial', size=10)

    # Geometría
    a = float (txt30.get())
    sheet['C16'] = a
    sheet['C16'].font = Font('Arial', size=10)

    b = float (txt31.get())
    sheet['C17'] = b
    sheet['C17'].font = Font('Arial', size=10)

    c = float (txt32.get())
    sheet['C18'] = c
    sheet['C18'].font = Font('Arial', size=10)

    h = float (txt33.get())
    sheet['C19'] = h
    sheet['D19'].font = Font('Arial', size=10)

    r = float (txt34.get())
    sheet['C20'] = r
    sheet['C20'].font = Font('Arial', size=10)

    bw = float (txt35.get())
    sheet['C21'] = bw
    sheet['C21'].font = Font('Arial', size=10)

    d = h-r

    # Cargas de diseño
    Vu = float (txt02.get())
    sheet['C24'] = Vu
    sheet['C24'].font = Font('Arial', size=10)

    Nuc1 = float (txt04.get())
    sheet['C25'] = Nuc1
    sheet['C25'].font = Font('Arial', size=10)

    Nuc2 = 0.2*Vu
    sheet['C26'] = "%.2f" %Nuc2
    sheet['C26'].font = Font('Arial', size=10)

    Nuc = max(Nuc1, Nuc2)
    sheet['C27'] = "%.2f" %Nuc
    sheet['C27'].font = Font('Arial', size=10)

    Mu = (Vu*a) + Nuc*(h-d)
    sheet['C28'] = "%.1f" %Mu
    sheet['C28'].font = Font('Arial', size=10)

    # Chequeos
    x = a/d
    if (x < 1):
        opc1 = 'Ok'
        sheet['C31'] = opc1
        sheet['C31'].font = Font('Arial', size=10)

    else:
        opc2 = 'No cumple'
        sheet['C31'] = opc2
        sheet['C31'].font = Font('Arial', size=10)
    
    hmin = 0.5*d
    sheet['C32'] = "%.2f" %hmin
    sheet['C32'].font = Font('Arial', size=10)

    if (Nuc1 <= Vu):
        opc1 = 'Ok'
        sheet['C33'] = opc1
        sheet['C33'].font = Font('Arial', size=10)
    else:
        opc2 = 'No cumple'
        sheet['C33'] = opc2
        sheet['C33'].font = Font('Arial', size=10)

    # Verificación Vu
    Vc1 = 0.2*fiv*fc*bw*d
    sheet['C36'] = "%.0f" %Vc1
    sheet['C36'].font = Font('Arial', size=10)

    Vc2 = ((3.3+0.08*fc*10)*fiv*bw*10*d*10)/1000
    sheet['C37'] = "%.0f" %Vc2
    sheet['C37'].font = Font('Arial', size=10)

    Vc3 = 11*fiv*bw*d
    sheet['C38'] = "%.0f" %Vc3
    sheet['C38'].font = Font('Arial', size=10)

    Vc = min(Vc1, Vc2, Vc3)
    sheet['C39'] = "%.0f" %Vc
    sheet['C39'].font = Font('Arial', size=10)

    if (Vu <= Vc):
        opc1 = 'Ok'
        sheet['C40'] = opc1
        sheet['C40'].font = Font('Arial', size=10)           
    else:
        opc2 = 'No cumple'
        sheet['C40'] = opc2    
        sheet['C40'].font = Font('Arial', size=10)
 
    # Refuerzo
    Avf = Vu/(fiv * fy * u)
    sheet['C50'] = "%.2f" %Avf
    sheet['C50'].font = Font('Arial', size=10)

    An = Nuc/(fif*fy)
    sheet['C52'] = "%.2f" %An
    sheet['C52'].font = Font('Arial', size=10)

    Af = Mu/(fif*fy*(d-a/2))
    sheet['C54'] = "%.2f" %Af
    sheet['C54'].font = Font('Arial', size=10)

    As1 = Af + An
    sheet['C56'] = "%.2f" %As1
    sheet['C56'].font = Font('Arial', size=10)

    As2 = 2/3*Avf + An
    sheet['C57'] = "%.2f" %As2
    sheet['C57'].font = Font('Arial', size=10)

    As3 = 0.04*(fc/fy)*bw*d
    sheet['C58'] = "%.2f" %As3
    sheet['C58'].font = Font('Arial', size=10)

    As = max(As1, As2, As3)
    sheet['C59'] = "%.2f" %As
    sheet['C59'].font = Font('Arial', size=10)

    # Refuerzo en estribos
    lde = 2/3*d    
    sheet['C63'] = "%.2f" %lde
    sheet['C63'].font = Font('Arial', size=10)

    Ah = 0.5*(As - An)    
    sheet['C64'] = "%.2f" %Ah
    sheet['C64'].font = Font('Arial', size=10)

    # Refuerzo Auxiliar
    Aa = 0.002*bw*c 
    sheet['C67'] = "%.2f" %Aa
    sheet['C67'].font = Font('Arial', size=10)

    # Verificación refuerzo mínimo
    import f_Asb as Asb
    Asb = Asb.f_Asb(Nb)
    import math
    Cant = math.ceil(As/Asb)
    Asc = Asb * Cant
    s = round((bw-15)/(Cant-1),0)
    sheet['C43'] = "%.2f" %Asc
    sheet['C43'].font = Font('Arial', size=10)

    pcal = Asc/(bw*d)
    sheet['C44'] = "%.5f" %pcal
    sheet['C44'].font = Font('Arial', size=10)

    pmin = 0.04*fc/fy
    sheet['C45'] = "%.5f" %pmin
    sheet['C45'].font = Font('Arial', size=10)
    
    if (pcal >= pmin):
        opc1 = 'Ok'
        sheet['C46'] = opc1
        sheet['C46'].font = Font('Arial', size=10)
    else:
        opc2 = 'No cumple'
        sheet['C46'] = opc2
        sheet['C46'].font = Font('Arial', size=10)

    # Distribución del refuerzo
    Reffp = str(Cant) + ' barras N° ' +  str("{:.0f}".format(Nb)) + " C/ " + "{:.1f}".format(s) + ' cm'    
    sheet['C70'] = Reffp
    sheet['C70'].font = Font('Arial', size=10)

    import f_Asb as Asb
    Asbe = Asb.f_Asb(Nbe)
    import math
    Cante = math.ceil(Ah/Asbe)
    se = round((lde/Cante), 0)
    Reffe = str(Cante) + ' estribos N° ' +  str("{:.0f}".format(Nbe)) + " C/ " + "{:.1f}".format(se) + ' cm'    
    sheet['C71'] = Reffe
    sheet['C71'].font = Font('Arial', size=10)

    import f_Asb as Asb
    AsbAa = Asb.f_Asb(Nb)
    CantAa = math.ceil(Aa/AsbAa)
    sAa = round((bw-15)/(CantAa-1),0)   
    Reffa = str(CantAa) + ' barras N° ' +  str("{:.0f}".format(Nb)) + " C/ " + "{:.1f}".format(sAa) + ' cm'    
    sheet['C72'] = Reffa
    sheet['C72'].font = Font('Arial', size=10)

    # Detalle soldadura
    import f_db as db
    db = db.f_db (Nb)
    sheet['C75'] = "%.2f" %db
    sheet['C75'].font = Font('Arial', size=10)

    lw = 3/4*db
    sheet['C76'] = "%.2f" %lw
    sheet['C76'].font = Font('Arial', size=10)

    tw = 1/2*db
    sheet['C77'] = "%.2f" %tw
    sheet['C77'].font = Font('Arial', size=10)

    #book.save ('D:\BIBLIOTECA PERSONAL\Programación\Python\Mensula\Diseño estructural ménsula.xlsx')         
    book.save ('D:\IEB_REPORTES\Diseño estructural ménsula.xlsx')
    
# Función para programar boton "Borrar"

def borrar():
    txt02.delete(0, 'end')   
    txt04.delete(0, 'end')   
    txt06.delete(0, 'end')   
    txt08.delete(0, 'end')   
    txt10.delete(0, 'end')
    txt21.delete(0, 'end')
    txt22.delete(0, 'end')
    txt23.delete(0, 'end')
    txt24.delete(0, 'end')
    txt25.delete(0, 'end')
    txt26.delete(0, 'end')
    txt27.delete(0, 'end')   
    txt30.delete(0, 'end')
    txt31.delete(0, 'end')
    txt32.delete(0, 'end')
    txt33.delete(0, 'end')
    txt34.delete(0, 'end')
    txt35.delete(0, 'end')
    rtxt2.delete(0, 'end')
    rtxt3.delete(0, 'end')   
    rtxt4.delete(0, 'end')   
    rtxt6.delete(0, 'end')   
    rtxt7.delete(0, 'end')   
    rtxt8.delete(0, 'end')   
    rtxt9.delete(0, 'end')   
    rtxt10.delete(0, 'end')   
    rtxt12.delete(0, 'end')   
    rtxt13.delete(0, 'end')   
    rtxt14.delete(0, 'end')   
    rtxt15.delete(0, 'end')
    rtxt22.delete(0, 'end')   
    rtxt24.delete(0, 'end')   
    rtxt25.delete(0, 'end')   
    rtxt27.delete(0, 'end')   
    rtxt28.delete(0, 'end')   
    rtxt29.delete(0, 'end')   
    rtxt30.delete(0, 'end')
    rtxt33.delete(0, 'end')
    rtxt34.delete(0, 'end')
    rtxt33.delete(0, 'end')
    dtxt1.delete(0, 'end')
    dtxt2.delete(0, 'end')
    dtxt3.delete(0, 'end')
    dtxt4.delete(0, 'end')
    dtxt5.delete(0, 'end')
    dtxt6.delete(0, 'end')

# Aqui comineza el recuadro 1

frame2 = LabelFrame(ventana, text = '  1 - Datos iniciales.  ')
frame2.pack()
frame2.place(x=20, y=10, width=375, height=360)

lb20 = Label(frame2, text = "Materiales", font='Helvetica 8 bold')
lb20.grid(row=0, column=0)

lb21 = Label(frame2, text = "   f'c. Resistencia del concreto _ MPa   ")
lb21.grid(row=1, column=0)
txt21 = Entry(frame2, justify=CENTER)
txt21.grid(row=1, column=1)

lb22 = Label(frame2, text = "fy. Fluencia del acero _ MPa")
lb22.grid(row=2, column=0)
txt22 = Entry(frame2, justify=CENTER)
txt22.grid(row=2, column=1)

lb23 = Label(frame2, text = "µ. Coeficiente de fricción")
lb23.grid(row=3, column=0)
txt23 = Entry(frame2, justify=CENTER)
txt23.grid(row=3, column=1)

lb24 = Label(frame2, text = "øv. Factor de reducción a cortante")
lb24.grid(row=4, column=0)
txt24 = Entry(frame2, justify=CENTER)
txt24.grid(row=4, column=1)

lb25 = Label(frame2, text = "øf. Factor de reducción a flexión")
lb25.grid(row=5, column=0)
txt25 = Entry(frame2, justify=CENTER)
txt25.grid(row=5, column=1)

lb26 = Label(frame2, text = "   N°. Número de la barra refuerzo principal   ")
lb26.grid(row=6, column=0)
txt26 = Entry(frame2, justify=CENTER)
txt26.grid(row=6, column=1)

lb27 = Label(frame2, text = "N°. Número de la barra estribos cerrados")
lb27.grid(row=7, column=0)
txt27 = Entry(frame2, justify=CENTER)
txt27.grid(row=7, column=1)

lb28 = Label(frame2, text = "")
lb28.grid(row=8, column=0)

lb29 = Label(frame2, text = "Geometría", font='Helvetica 8 bold')
lb29.grid(row=9, column=0)

lb30 = Label(frame2, text = "av. Dist. aplicación de carga _ cm")
lb30.grid(row=10, column=0)
txt30 = Entry(frame2, justify=CENTER)
txt30.grid(row=10, column=1)

lb31 = Label(frame2, text = "b. Ancho del soporte _ cm")
lb31.grid(row=11, column=0)
txt31 = Entry(frame2, justify=CENTER)
txt31.grid(row=11, column=1)

lb32 = Label(frame2, text = "c. Ancho de la ménsula _ cm")
lb32.grid(row=12, column=0)
txt32 = Entry(frame2, justify=CENTER)
txt32.grid(row=12, column=1)

lb33 = Label(frame2, text = "h. Altura de la ménsula _ cm")
lb33.grid(row=13, column=0)
txt33 = Entry(frame2, justify=CENTER)
txt33.grid(row=13, column=1)

lb34 = Label(frame2, text = "r. Recubrimiento del refuerzo _ cm")
lb34.grid(row=14, column=0)
txt34 = Entry(frame2, justify=CENTER)
txt34.grid(row=14, column=1)

lb35 = Label(frame2, text = "bw. Largo de la ménsula _ cm")
lb35.grid(row=15, column=0)
txt35 = Entry(frame2, justify=CENTER)
txt35.grid(row=15, column=1)

# Aqui comineza el recuadro 2

frame1 = LabelFrame(ventana, text = '  2 - Cargas de diseño.  ')
frame1.pack()
frame1.place(x=405, y=10, width=375, height=360)

lb01 = Label(frame1, text = "")
lb01.grid(row=0, column=0)

lb02 = Label(frame1, text = "Vu. Fuerza cortante _ kN")
lb02.grid(row=1, column=0)
txt02 = Entry(frame1, justify=CENTER)
txt02.grid(row=1, column=1)

lb03 = Label(frame1, text = "")
lb03.grid(row=2, column=0)

lb04 = Label(frame1, text = "Nuc. Fuerza horizontal _ kN")
lb04.grid(row=3, column=0)
txt04 = Entry(frame1, justify=CENTER)
txt04.grid(row=3, column=1)

lb05 = Label(frame1, text = "")
lb05.grid(row=4, column=0)

lb06 = Label(frame1, text = "   Nuc cal = 0.2 * Vu _ kN   ")
lb06.grid(row=5, column=0)
txt06 = Entry(frame1, justify=CENTER)
txt06.grid(row=5, column=1)
txt06.config(bg = "#ecf0f1")

lb07 = Label(frame1, text = "")
lb07.grid(row=6, column=0)

lb08 = Label(frame1, text = "Nuc = máx(Nuc, Nuc cal)")
lb08.grid(row=7, column=0)
txt08 = Entry(frame1, justify=CENTER)
txt08.grid(row=7, column=1)
txt08.config(bg = "#ecf0f1")

lb09 = Label(frame1, text = "")
lb09.grid(row=8, column=0)

lb10 = Label(frame1, text = "    Mu. Momento flector _ kN.cm    ")
lb10.grid(row=9, column=0)
txt10 = Entry(frame1, justify=CENTER)
txt10.grid(row=9, column=1)
txt10.config(bg = "#ecf0f1")

# Aqui comineza el recuadro 3

frame3 = LabelFrame(ventana, text = '  3 - Resultados.  ')
frame3.pack()
frame3.place(x=20, y=370, width=760, height=430)

rlb1 = Label(frame3, text = "Chequeos", font='Helvetica 8 bold')
rlb1.grid(row=0, column=0)

elb1 = Label(frame3, text = " ")
elb1.grid(row=1, column=0)

rlb2 = Label(frame3, text = "Relación a/d < 1.0")
rlb2.grid(row=2, column=0)
rtxt2 = Entry(frame3, justify=CENTER)
rtxt2.grid(row=2, column=1)
rtxt2.config(bg = "#ecf0f1")

rlb3 = Label(frame3, text = "hm: hmín. = 0.5 * d _ cm")
rlb3.grid(row=3, column=0)
rtxt3 = Entry(frame3, justify=CENTER)
rtxt3.grid(row=3, column=1)
rtxt3.config(bg = "#ecf0f1")

rlb4 = Label(frame3, text = "Nuc ≤ Vu")
rlb4.grid(row=4, column=0)
rtxt4 = Entry(frame3, justify=CENTER)
rtxt4.grid(row=4, column=1)
rtxt4.config(bg = "#ecf0f1")

elb2 = Label(frame3, text = " ")
elb2.grid(row=5, column=0)

rlb5 = Label(frame3, text = "Verificación Vu _ kN")
rlb5.grid(row=6, column=0)

rlb6 = Label(frame3, text = "Vc1 = 0.2 * ø * f'c * bw * d")
rlb6.grid(row=7, column=0)
rtxt6 = Entry(frame3, justify=CENTER)
rtxt6.grid(row=7, column=1)
rtxt6.config(bg = "#ecf0f1")

rlb7 = Label(frame3, text = "   Vc2 = (3.3 + 0.08 * f'c) * ø * bw * d   ")
rlb7.grid(row=8, column=0)
rtxt7 = Entry(frame3, justify=CENTER)
rtxt7.grid(row=8, column=1)
rtxt7.config(bg = "#ecf0f1")

rlb8 = Label(frame3, text = "Vc3 = 11 * ø * bw * d")
rlb8.grid(row=9, column=0)
rtxt8 = Entry(frame3, justify=CENTER)
rtxt8.grid(row=9, column=1)
rtxt8.config(bg = "#ecf0f1")

rlb9 = Label(frame3, text = "Vc = mín. (Vc1, Vc2, Vc3)")
rlb9.grid(row=10, column=0)
rtxt9 = Entry(frame3, justify=CENTER)
rtxt9.grid(row=10, column=1)
rtxt9.config(bg = "#ecf0f1")

rlb10 = Label(frame3, text = "Vu ≤ Vc")
rlb10.grid(row=11, column=0)
rtxt10 = Entry(frame3, justify=CENTER)
rtxt10.grid(row=11, column=1)
rtxt10.config(bg = "#ecf0f1")

elb3 = Label(frame3, text = " ")
elb3.grid(row=12, column=0)

rlb11 = Label(frame3, text = "Verificación refuerzo mínimo")
rlb11.grid(row=13, column=0)

rlb12 = Label(frame3, text = "Asc. Acero colocado _ cm²")
rlb12.grid(row=14, column=0)
rtxt12 = Entry(frame3, justify=CENTER)
rtxt12.grid(row=14, column=1)
rtxt12.config(bg = "#ecf0f1")

rlb13 = Label(frame3, text = "pcal = Asc / (bw * d)")
rlb13.grid(row=15, column=0)
rtxt13 = Entry(frame3, justify=CENTER)
rtxt13.grid(row=15, column=1)
rtxt13.config(bg = "#ecf0f1")

rlb14 = Label(frame3, text = "pmín = 0.04 * f'c / fy")
rlb14.grid(row=16, column=0)
rtxt14 = Entry(frame3, justify=CENTER)
rtxt14.grid(row=16, column=1)
rtxt14.config(bg = "#ecf0f1")

rlb15 = Label(frame3, text = "pcal ≥ pmín")
rlb15.grid(row=17, column=0)
rtxt15 = Entry(frame3, justify=CENTER)
rtxt15.grid(row=17, column=1)
rtxt15.config(bg = "#ecf0f1")

rlb20 = Label(frame3, text = "Refuerzo", font='Helvetica 8 bold')
rlb20.grid(row=0, column=2)

elb4 = Label(frame3, text = " ")
elb4.grid(row=1, column=2)

rlb21 = Label(frame3, text = "Refuerzo para resistir Vu. _ cm²")
rlb21.grid(row=2, column=2)

rlb22 = Label(frame3, text = "Avf = Vu / (ø * fy * μ)")
rlb22.grid(row=3, column=2)
rtxt22 = Entry(frame3, justify=CENTER)
rtxt22.grid(row=3, column=3)
rtxt22.config(bg = "#ecf0f1")

rlb23 = Label(frame3, text = "     Refuerzo para resistir Nuc. _ cm²     ")
rlb23.grid(row=4, column=2)

rlb24 = Label(frame3, text = "An = Nuc / (ø * fy)")
rlb24.grid(row=5, column=2)
rtxt24 = Entry(frame3, justify=CENTER)
rtxt24.grid(row=5, column=3)
rtxt24.config(bg = "#ecf0f1")

rlb25 = Label(frame3, text = "Refuerzo para resistir Mu. _ cm²")
rlb25.grid(row=6, column=2)

rlb25 = Label(frame3, text = "Af = Mu / (ø * fy * (d - a / 2))")
rlb25.grid(row=7, column=2)
rtxt25 = Entry(frame3, justify=CENTER)
rtxt25.grid(row=7, column=3)
rtxt25.config(bg = "#ecf0f1")

elb5 = Label(frame3, text = " ")
elb5.grid(row=8, column=2)

rlb26 = Label(frame3, text = "Refuerzo principal _ cm²")
rlb26.grid(row=9, column=2)

rlb27 = Label(frame3, text = "As1 = Af + An")
rlb27.grid(row=10, column=2)
rtxt27 = Entry(frame3, justify=CENTER)
rtxt27.grid(row=10, column=3)
rtxt27.config(bg = "#ecf0f1")

rlb28 = Label(frame3, text = "As2 = 2/3 * Avf + An")
rlb28.grid(row=11, column=2)
rtxt28 = Entry(frame3, justify=CENTER)
rtxt28.grid(row=11, column=3)
rtxt28.config(bg = "#ecf0f1")

rlb29 = Label(frame3, text = "As3 = 0.04 * (f'c / fy ) * bw * d")
rlb29.grid(row=12, column=2)
rtxt29 = Entry(frame3, justify=CENTER)
rtxt29.grid(row=12, column=3)
rtxt29.config(bg = "#ecf0f1")

rlb30 = Label(frame3, text = "As = máx (As1, As2, As3)")
rlb30.grid(row=13, column=2)
rtxt30 = Entry(frame3, justify=CENTER)
rtxt30.grid(row=13, column=3)
rtxt30.config(bg = "#ecf0f1")

elb6 = Label(frame3, text = " ")
elb6.grid(row=14, column=2)

rlb31 = Label(frame3, text = "Refuerzo en estribos _ cm²")
rlb31.grid(row=15, column=2)

rlb32 = Label(frame3, text = "Espacio para distribuir estribos")
rlb32.grid(row=16, column=2)

rlb33 = Label(frame3, text = "2/3 * d _ cm")
rlb33.grid(row=17, column=2)
rtxt33 = Entry(frame3, justify=CENTER)
rtxt33.grid(row=17, column=3)
rtxt33.config(bg = "#ecf0f1")

rlb34 = Label(frame3, text = "Ah = 0.5*(As-An)")
rlb34.grid(row=18, column=2)
rtxt34 = Entry(frame3, justify=CENTER)
rtxt34.grid(row=18, column=3)
rtxt34.config(bg = "#ecf0f1")

# Aqui comineza el recuadro 4

frame4 = LabelFrame(ventana, text = '  4 - Distribución del refuerzo y detalle soldadura.  ')
frame4.pack()
frame4.place(x=20, y=800, width=760, height=150)

dlb1 = Label(frame4, text = "Distribución refuerzo principal")
dlb1.place(x=20, y=15, width=160, height=20)
dtxt1 = Entry(frame4, justify=CENTER)
dtxt1.place(x=190, y=15, width=150, height=20)
dtxt1.config(bg = "#ecf0f1")

dlb2 = Label(frame4, text = "Distribución estribos cerrados")
dlb2.place(x=20, y=50, width=160, height=20)
dtxt2 = Entry(frame4, justify=CENTER)
dtxt2.place(x=190, y=50, width=150, height=20)
dtxt2.config(bg = "#ecf0f1")

dlb3 = Label(frame4, text = "Distribución refuerzo auxiliar")
dlb3.place(x=20, y=85, width=160, height=20)
dtxt3 = Entry(frame4, justify=CENTER)
dtxt3.place(x=190, y=85, width=150, height=20)
dtxt3.config(bg = "#ecf0f1")

dlb4 = Label(frame4, text = "db. Diámetro refuerzo principal _ cm")
dlb4.place(x=360, y=15, width=200, height=20)
dtxt4 = Entry(frame4, justify=CENTER)
dtxt4.place(x=570, y=15, width=80, height=20)
dtxt4.config(bg = "#ecf0f1")

dlb5 = Label(frame4, text = "lw. Long de la soldadura _ cm")
dlb5.place(x=360, y=50, width=200, height=20)
dtxt5 = Entry(frame4, justify=CENTER)
dtxt5.place(x=570, y=50, width=80, height=20)
dtxt5.config(bg = "#ecf0f1")

dlb6 = Label(frame4, text = "tw. Long de la soldadura _ cm")
dlb6.place(x=360, y=85, width=200, height=20)
dtxt6 = Entry(frame4, justify=CENTER)
dtxt6.place(x=570, y=85, width=80, height=20)
dtxt6.config(bg = "#ecf0f1")

# Aqui comineza el recuadro 5

frame5 = LabelFrame(ventana, text = '  5 - Diagramas representativos.  ')
frame5.pack()
frame5.place(x=790, y=10, width=390, height=940)

from PIL import ImageTk, Image

#img = Image.open('D:\BIBLIOTECA PERSONAL\Programación\Python\Mensula\Diagramas.jpg')
img = Image.open('D:\IEB_MENSULA\Diagramas.jpg')
img = img.resize((380,907), Image.ANTIALIAS)

img = ImageTk.PhotoImage(img)
lbl_img = Label(frame5, image=img)
lbl_img.pack()

# Aqui comienzan los botones

boton1 = Button(ventana, text = 'Calcular', command=calcular)
boton1.pack()
boton1.place(x=320, y=960, width=150, height=20)

boton2 = Button(ventana, text = 'Guardar en xls', command=reporte)
boton2.pack()
boton2.place(x=520, y=960, width=150, height=20)

boton3 = Button(ventana, text = 'Borrar', command=borrar)
boton3.pack()
boton3.place(x=720, y=960, width=150, height=20)

label = Label(ventana, text = "wilson.taimalc@gmail.com - 2023", font='Arial 7')
label.pack()
label.place(x=955, y=970, width=225, height=10)

ventana.mainloop ()
